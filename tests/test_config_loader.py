from __future__ import annotations

import json
import os
import sys
import tempfile
import unittest
from pathlib import Path
from unittest.mock import patch

from openpyxl import Workbook

ROOT = Path(__file__).resolve().parent.parent
if str(ROOT) not in sys.path:
    sys.path.insert(0, str(ROOT))

from config_loader import ConfigError, load_probe_config_from_dict, map_success_mode
from probe_core import (
    SUCCESS_MODE_ANY_HTTP,
    SUCCESS_MODE_CUSTOM,
    column_letter_to_index,
    resolve_url_column_index,
)


class ConfigLoaderTests(unittest.TestCase):
    def setUp(self) -> None:
        self.temp_dir = tempfile.TemporaryDirectory()
        self.addCleanup(self.temp_dir.cleanup)
        self.workbook_path = os.path.join(self.temp_dir.name, "sample.xlsx")
        wb = Workbook()
        ws = wb.active
        ws.title = "Sheet1"
        ws["A1"] = "名称"
        ws["B1"] = "网址"
        ws["A2"] = "示例"
        ws["B2"] = "https://example.com"
        wb.save(self.workbook_path)

    def _base_config(self, **probe_overrides: object) -> dict:
        probe = {
            "sheet_name": "Sheet1",
            "has_header": True,
            "url_column": "B",
            "concurrency": 5,
            "timeout_seconds": 30,
            "retry_count": 1,
            "success_mode": "any_http",
        }
        probe.update(probe_overrides)
        return {
            "input": {
                "file_path": self.workbook_path,
                "password": "",
                "use_office_fallback": False,
            },
            "probe": probe,
            "output": {},
        }

    def test_load_defaults_and_output_path(self) -> None:
        config = load_probe_config_from_dict(self._base_config())
        self.assertEqual(config.url_col_idx, 2)
        self.assertEqual(config.success_mode, SUCCESS_MODE_ANY_HTTP)
        self.assertTrue(config.output_path.endswith("_monitoring_result.xlsx"))

    def test_success_mode_mapping(self) -> None:
        self.assertEqual(map_success_mode("2xx_3xx"), "仅 2xx/3xx")
        self.assertEqual(map_success_mode("custom"), SUCCESS_MODE_CUSTOM)
        with self.assertRaises(ConfigError):
            map_success_mode("invalid-mode")

    def test_url_column_letter_and_header_name(self) -> None:
        wb = Workbook()
        ws = wb.active
        ws.title = "Data"
        ws["A1"] = "网址"
        ws["A2"] = "https://example.com"
        path = os.path.join(self.temp_dir.name, "header.xlsx")
        wb.save(path)

        cfg_by_header = load_probe_config_from_dict(
            {
                "input": {"file_path": path, "password": "", "use_office_fallback": False},
                "probe": {
                    "sheet_name": "Data",
                    "has_header": True,
                    "url_column": "网址",
                    "success_mode": "any_http",
                },
                "output": {"path": os.path.join(self.temp_dir.name, "out.xlsx")},
            }
        )
        self.assertEqual(cfg_by_header.url_col_idx, 1)

        cfg_by_number = load_probe_config_from_dict(self._base_config(url_column=2))
        self.assertEqual(cfg_by_number.url_col_idx, 2)

    def test_password_env_priority(self) -> None:
        data = self._base_config()
        data["input"]["password_env"] = "TEST_EXCEL_PASSWORD"
        mock_wb = Workbook()
        mock_ws = mock_wb.active
        mock_ws.title = "Sheet1"
        mock_ws["B1"] = "网址"
        mock_ws["B2"] = "https://example.com"
        with patch.dict(os.environ, {"TEST_EXCEL_PASSWORD": "secret"}):
            with patch("config_loader.load_workbook_from_input", return_value=(mock_wb, None)):
                config = load_probe_config_from_dict(data)
        self.assertEqual(config.password, "secret")

    def test_custom_mode_requires_codes(self) -> None:
        with self.assertRaises(ConfigError):
            load_probe_config_from_dict(self._base_config(success_mode="custom"))

    def test_office_fallback_rejected_on_non_windows(self) -> None:
        data = self._base_config()
        data["input"]["use_office_fallback"] = True
        with patch.object(sys, "platform", "linux"):
            with self.assertRaises(ConfigError):
                load_probe_config_from_dict(data)

    def test_missing_file_raises(self) -> None:
        data = self._base_config()
        data["input"]["file_path"] = os.path.join(self.temp_dir.name, "missing.xlsx")
        with self.assertRaises(ConfigError):
            load_probe_config_from_dict(data)


class UrlColumnHelperTests(unittest.TestCase):
    def test_column_letter_to_index(self) -> None:
        self.assertEqual(column_letter_to_index("A"), 1)
        self.assertEqual(column_letter_to_index("B"), 2)
        self.assertEqual(column_letter_to_index("AA"), 27)

    def test_resolve_invalid_column(self) -> None:
        wb = Workbook()
        ws = wb.active
        ws.title = "Sheet1"
        with self.assertRaises(ValueError):
            resolve_url_column_index(ws, "missing-header", True)


if __name__ == "__main__":
    unittest.main()
