from __future__ import annotations

import asyncio
import io
import ipaddress
import json
import os
import re
import sys
import tempfile
import urllib.request
from dataclasses import dataclass
from datetime import datetime
from typing import Callable, Literal

from openpyxl import Workbook, load_workbook
from openpyxl.worksheet.worksheet import Worksheet
from playwright.async_api import BrowserContext, TimeoutError as PlaywrightTimeoutError
from playwright.async_api import async_playwright

try:
    import msoffcrypto  # type: ignore[attr-defined]
except Exception:
    msoffcrypto = None

try:
    import win32com.client as win32  # type: ignore
except Exception:
    win32 = None


DEFAULT_TIMEOUT_SECONDS = 30
DEFAULT_CONCURRENCY = 5
DEFAULT_RETRY_COUNT = 1

SUCCESS_MODE_2XX_3XX = "仅 2xx/3xx"
SUCCESS_MODE_2XX_3XX_OR_NO_STATUS = "2xx/3xx + 无状态码也算成功"
SUCCESS_MODE_ANY_HTTP = "2xx/3xx/4xx/5xx（任意状态码）"
SUCCESS_MODE_CUSTOM = "自定义 HTTP 状态码"


def get_app_base_dir() -> str:
    if getattr(sys, "frozen", False):
        return os.path.dirname(os.path.abspath(sys.executable))
    return os.path.dirname(os.path.abspath(__file__))


def get_runtime_temp_dir() -> str:
    temp_dir = os.path.join(get_app_base_dir(), "runtime-data", "temp")
    os.makedirs(temp_dir, exist_ok=True)
    return temp_dir


@dataclass
class ProbeResult:
    thread_name: str
    url: str
    accessible: bool
    detail: str
    status_code: int | None = None
    attempts: int = 1
    title: str | None = None
    redirect_count: int = 0
    final_url: str | None = None


@dataclass
class ProbeConfig:
    file_path: str
    use_office_fallback: bool
    sheet_name: str
    has_header: bool
    url_col_idx: int
    concurrency: int
    timeout_seconds: int
    retry_count: int
    success_mode: str
    custom_codes_input: str
    output_path: str
    password: str = ""
    probe_mode: str = "lightweight"
    extract_title: bool = True
    track_redirects: bool = True


@dataclass
class ExecuteProbeResult:
    outcome: Literal["success", "aborted"]
    success_count: int
    total_count: int
    failed_count: int
    output_path: str


def normalize_url(url_value: object) -> str:
    if url_value is None:
        return ""
    url = str(url_value).strip()
    if not url:
        return ""
    from urllib.parse import urlparse

    parsed = urlparse(url)
    if not parsed.scheme:
        return f"http://{url}"
    return url


def is_valid_url_or_domain(url_value: object) -> bool:
    if url_value is None:
        return False
    raw = str(url_value).strip()
    if not raw:
        return False

    from urllib.parse import urlparse

    candidate = raw if "://" in raw else f"http://{raw}"
    try:
        parsed = urlparse(candidate)
    except Exception:
        return False

    host = (parsed.hostname or "").strip()
    if not host:
        return False
    if host.lower() == "localhost":
        return True

    try:
        _: ipaddress.IPv4Address | ipaddress.IPv6Address = ipaddress.ip_address(host)
        return True
    except ValueError:
        pass

    domain_pattern = re.compile(r"^(?=.{1,253}$)([a-zA-Z0-9-]{1,63}\.)+[a-zA-Z]{2,63}$")
    return bool(domain_pattern.match(host))


def parse_custom_status_rules(raw_rules: str) -> tuple[set[int], set[int]]:
    exact_codes: set[int] = set()
    class_codes: set[int] = set()
    tokens = [token.strip() for token in raw_rules.replace("，", ",").split(",") if token.strip()]
    if not tokens:
        raise ValueError("empty")

    for token in tokens:
        lower = token.lower()
        if re.fullmatch(r"[1-5]xx", lower):
            class_codes.add(int(lower[0]))
            continue
        if re.fullmatch(r"\d{3}", token):
            code = int(token)
            if 100 <= code <= 599:
                exact_codes.add(code)
                continue
        raise ValueError(token)

    return exact_codes, class_codes


def build_success_checker(
    success_mode: str,
    custom_codes_input: str,
) -> Callable[[int | None, bool], bool]:
    custom_codes: set[int] = set()
    custom_code_classes: set[int] = set()
    if success_mode == SUCCESS_MODE_CUSTOM:
        try:
            custom_codes, custom_code_classes = parse_custom_status_rules(custom_codes_input)
        except ValueError as exc:
            raise RuntimeError(
                f"自定义状态码格式错误: {exc}。请按示例输入：200,301,302 或 2xx,3xx,5xx"
            ) from exc

    def success_checker(status_code: int | None, has_response: bool) -> bool:
        if success_mode == SUCCESS_MODE_2XX_3XX:
            return status_code is not None and 200 <= status_code < 400
        if success_mode == SUCCESS_MODE_2XX_3XX_OR_NO_STATUS:
            return (status_code is not None and 200 <= status_code < 400) or not has_response
        if success_mode == SUCCESS_MODE_ANY_HTTP:
            return status_code is not None
        return status_code is not None and (
            status_code in custom_codes or (status_code // 100) in custom_code_classes
        )

    return success_checker


def decrypt_with_password(file_path: str, password: str) -> io.BytesIO:
    if msoffcrypto is None:
        raise RuntimeError("未安装 msoffcrypto-tool，无法解密加密 Excel。")
    decrypted = io.BytesIO()
    with open(file_path, "rb") as f:
        office_file = msoffcrypto.OfficeFile(f)  # type: ignore
        office_file.load_key(password=password)  # type: ignore
        office_file.decrypt(decrypted)  # type: ignore
    _: int = decrypted.seek(0)
    return decrypted


def decrypt_with_office(file_path: str, password: str) -> io.BytesIO:
    if win32 is None:
        raise RuntimeError("当前环境不可用 pywin32，无法通过 Office/WPS 解密。")
    app = None
    workbook = None
    temp_output = tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False, dir=get_runtime_temp_dir())
    temp_output.close()
    prog_ids = ["Excel.Application", "ket.Application"]
    try:
        last_error: Exception | None = None
        for prog_id in prog_ids:
            try:
                app = win32.DispatchEx(prog_id)  # type: ignore
                break
            except Exception as exc:
                last_error = exc
                app = None
        if app is None:
            raise RuntimeError(f"无法启动 Office/WPS 自动化组件: {last_error}")
        app.Visible = False  # type: ignore
        app.DisplayAlerts = False  # type: ignore
        workbook = app.Workbooks.Open(file_path, False, True, None, password)  # type: ignore
        workbook.SaveAs(temp_output.name, FileFormat=51)  # type: ignore
        workbook.Close(False)  # type: ignore
        workbook = None
        app.Quit()  # type: ignore
        app = None
        with open(temp_output.name, "rb") as f:
            content = f.read()
        return io.BytesIO(content)
    finally:
        if workbook is not None:
            workbook.Close(False)  # type: ignore
        if app is not None:
            app.Quit()  # type: ignore
        if os.path.exists(temp_output.name):
            os.remove(temp_output.name)


def load_workbook_from_input(file_path: str, password: str, use_office_fallback: bool):
    if use_office_fallback and sys.platform != "win32":
        raise RuntimeError("Office/WPS 解密兜底仅在 Windows 上可用，请在 Linux 上使用 msoffcrypto 或关闭 use_office_fallback。")
    if not os.path.exists(file_path):
        raise FileNotFoundError(f"文件不存在: {file_path}")
    if password:
        try:
            decrypted = decrypt_with_password(file_path, password)
            wb = load_workbook(decrypted)
            return wb, decrypted
        except Exception as first_error:
            if not use_office_fallback:
                raise RuntimeError(f"密码解密失败: {first_error}") from first_error
            decrypted = decrypt_with_office(file_path, password)
            wb = load_workbook(decrypted)
            return wb, decrypted
    wb = load_workbook(file_path)
    return wb, None


def column_letter_to_index(column: str) -> int:
    value = column.strip().upper()
    if not value or not re.fullmatch(r"[A-Z]+", value):
        raise ValueError(f"无效的列字母: {column!r}")
    index = 0
    for ch in value:
        index = index * 26 + (ord(ch) - ord("A") + 1)
    return index


def resolve_url_column_index(sheet: Worksheet, url_column: object, has_header: bool) -> int:
    if isinstance(url_column, bool):
        raise ValueError("url_column 不能为布尔值")
    if isinstance(url_column, int):
        if url_column < 1:
            raise ValueError("url_column 列号必须 >= 1")
        return url_column
    if isinstance(url_column, float) and url_column.is_integer():
        col_idx = int(url_column)
        if col_idx < 1:
            raise ValueError("url_column 列号必须 >= 1")
        return col_idx

    text = str(url_column).strip()
    if not text:
        raise ValueError("url_column 不能为空")
    if re.fullmatch(r"\d+", text):
        col_idx = int(text)
        if col_idx < 1:
            raise ValueError("url_column 列号必须 >= 1")
        return col_idx
    if re.fullmatch(r"[A-Za-z]+", text):
        return column_letter_to_index(text)
    if has_header:
        for col_idx in range(1, sheet.max_column + 1):
            header_value = sheet.cell(row=1, column=col_idx).value
            if header_value is not None and str(header_value).strip() == text:
                return col_idx
        raise ValueError(f"未在表头中找到列名: {text!r}")
    raise ValueError(f"无法解析 url_column: {text!r}（无表头时需使用列号或列字母）")


def extract_urls_from_sheet(sheet: Worksheet, url_col_idx: int, has_header: bool) -> list[str]:
    start_row = 2 if has_header else 1
    urls: list[str] = []
    for row_idx in range(start_row, sheet.max_row + 1):
        raw = sheet.cell(row=row_idx, column=url_col_idx).value
        urls.append(normalize_url(raw))
    return urls


def localize_isp_name(raw_isp: str) -> str:
    isp = (raw_isp or "").strip()
    if not isp:
        return "未知运营商"
    if any("\u4e00" <= ch <= "\u9fff" for ch in isp):
        return isp

    text = isp.lower()
    province_map = {
        "beijing": "北京",
        "tianjin": "天津",
        "hebei": "河北",
        "shanxi": "山西",
        "inner mongolia": "内蒙古",
        "liaoning": "辽宁",
        "jilin": "吉林",
        "heilongjiang": "黑龙江",
        "shanghai": "上海",
        "jiangsu": "江苏",
        "zhejiang": "浙江",
        "anhui": "安徽",
        "fujian": "福建",
        "jiangxi": "江西",
        "shandong": "山东",
        "henan": "河南",
        "hubei": "湖北",
        "hunan": "湖南",
        "guangdong": "广东",
        "guangxi": "广西",
        "hainan": "海南",
        "chongqing": "重庆",
        "sichuan": "四川",
        "guizhou": "贵州",
        "yunnan": "云南",
        "tibet": "西藏",
        "shaanxi": "陕西",
        "gansu": "甘肃",
        "qinghai": "青海",
        "ningxia": "宁夏",
        "xinjiang": "新疆",
        "hong kong": "香港",
        "macau": "澳门",
        "taiwan": "台湾",
    }

    province = ""
    for key, value in province_map.items():
        if key in text:
            province = value
            break

    carrier = ""
    if "china unicom" in text:
        carrier = "中国联通"
    elif "china telecom" in text:
        carrier = "中国电信"
    elif "china mobile" in text:
        carrier = "中国移动"
    elif "cernet" in text or "education and research network" in text:
        carrier = "中国教育网"
    elif "alibaba" in text or "aliyun" in text:
        carrier = "阿里云"
    elif "tencent" in text:
        carrier = "腾讯云"
    elif "huawei cloud" in text or "huaweicloud" in text:
        carrier = "华为云"
    elif "baidu" in text:
        carrier = "百度云"

    if carrier and province:
        return f"{carrier}（{province}）"
    if carrier:
        return carrier
    return isp


def detect_public_ip_and_isp(timeout_seconds: int = 8) -> tuple[str, str, str]:
    def fetch_text(url: str) -> str:
        req = urllib.request.Request(url, headers={"User-Agent": "SyntheticMonitoringTool/1.0"})
        with urllib.request.urlopen(req, timeout=timeout_seconds) as resp:  # type: ignore
            return resp.read().decode("utf-8", errors="replace").strip()  # type: ignore

    def fetch_json(url: str) -> dict[str, str]:
        return json.loads(fetch_text(url))  # type: ignore

    try:
        ipip_response = fetch_text("https://myip.ipip.net")
        ip_match = re.search(r"IP[：:]\s*(\d+\.\d+\.\d+\.\d+)", ipip_response)
        ip = ip_match.group(1) if ip_match else ""
        location_match = re.search(r"来自于[：:]\s*(.+)", ipip_response)
        if location_match:
            location_str = location_match.group(1).strip()
            location_parts = location_str.split()
            isp = location_parts[-1] if location_parts else ""
            geo_parts = location_parts[:-1] if len(location_parts) > 1 else []
            location = "/".join(geo_parts) if geo_parts else "未知地区"
        else:
            isp = ""
            location = "未知地区"
        if ip:
            return ip, localize_isp_name(isp), location
    except Exception:
        pass

    def is_valid_ipv4(ip_str: str) -> bool:
        try:
            addr = ipaddress.ip_address(ip_str)
            return isinstance(addr, ipaddress.IPv4Address)
        except ValueError:
            return False

    ip_candidates = [
        ("ipify", lambda: str(fetch_json("https://api.ipify.org?format=json").get("ip", "")).strip()),  # type: ignore
        ("ip.sb", lambda: str(fetch_json("https://api-ipv4.ip.sb/geoip").get("ip", "")).strip()),  # type: ignore
        ("icanhazip", lambda: fetch_text("https://ipv4.icanhazip.com").splitlines()[0].strip()),
    ]

    ip = ""
    for _, getter in ip_candidates:
        try:
            value = getter()
            if value and is_valid_ipv4(value):
                ip = value
                break
        except Exception:
            continue

    if not ip:
        return "检测失败", "检测失败（网络受限）", "检测失败"

    geo_sources = [
        ("ipwho.is", lambda: fetch_json(f"https://ipwho.is/{ip}?lang=zh")),
        ("ipapi.co", lambda: fetch_json(f"https://ipapi.co/{ip}/json/")),
    ]

    for source_name, getter in geo_sources:
        try:
            data = getter()
            if source_name == "ipwho.is":
                if not bool(data.get("success", False)):  # type: ignore
                    continue
                connection: dict[str, object] = data.get("connection") or {}  # type: ignore
                isp = str(connection.get("isp") or connection.get("org") or "").strip()  # type: ignore
                country = str(data.get("country", "")).strip()  # type: ignore
                region = str(data.get("region", "")).strip()  # type: ignore
                city = str(data.get("city", "")).strip()  # type: ignore
            else:
                if str(data.get("error", "")).strip():  # type: ignore
                    continue
                isp = str(data.get("org", "")).strip()  # type: ignore
                country = str(data.get("country_name", "")).strip()  # type: ignore
                region = str(data.get("region", "")).strip()  # type: ignore
                city = str(data.get("city", "")).strip()  # type: ignore
            location_parts = [part for part in [country, region, city] if part]
            location = "/".join(location_parts) if location_parts else "未知地区"
            return ip, localize_isp_name(isp), location
        except Exception:
            continue

    return ip, "未知运营商", "未知地区"


async def probe_url(
    context: BrowserContext,
    worker_name: str,
    url: str,
    timeout_ms: int,
    success_checker: Callable[[int | None, bool], bool],
    retry_count: int,
) -> ProbeResult:
    if not url:
        return ProbeResult(thread_name=worker_name, url="", accessible=False, detail="空链接", attempts=1)
    last_status: int | None = None
    last_detail = "未知错误"
    attempts_total = retry_count + 1
    for attempt in range(1, attempts_total + 1):
        page = await context.new_page()
        try:
            response = await page.goto(url, wait_until="domcontentloaded", timeout=timeout_ms)
            has_response = response is not None
            raw_status = response.status if response is not None else None
            if raw_status is not None and isinstance(raw_status, int):
                last_status = raw_status
                last_detail = f"HTTP {raw_status}"
            else:
                last_status = None
                last_detail = "加载成功(无状态码)" if has_response else "无响应"
            if success_checker(last_status, has_response):
                detail = last_detail if attempt == 1 else f"{last_detail}，重试后成功({attempt}/{attempts_total})"
                return ProbeResult(
                    thread_name=worker_name,
                    url=url,
                    accessible=True,
                    detail=detail,
                    status_code=last_status,
                    attempts=attempt,
                )
        except PlaywrightTimeoutError:
            last_status = None
            last_detail = "超时"
        except Exception as exc:
            last_status = None
            last_detail = str(exc)
        finally:
            await page.close()
    detail = f"{last_detail}，已重试 {retry_count} 次" if retry_count > 0 else last_detail
    return ProbeResult(
        thread_name=worker_name,
        url=url,
        accessible=False,
        detail=detail,
        status_code=last_status,
        attempts=attempts_total,
    )


async def probe_url_httpx(
    worker_name: str,
    url: str,
    timeout_ms: int,
    success_checker: Callable[[int | None, bool], bool],
    retry_count: int,
    extract_title: bool,
    track_redirects: bool,
) -> ProbeResult:
    """轻量拨测：httpx + BeautifulSoup，无浏览器开销。"""
    import httpx  # type: ignore
    from bs4 import BeautifulSoup  # type: ignore

    if not url:
        return ProbeResult(thread_name=worker_name, url="", accessible=False, detail="空链接", attempts=1)

    last_status: int | None = None
    last_detail = "未知错误"
    last_title: str | None = None
    last_redirect_count = 0
    last_final_url: str | None = None
    attempts_total = retry_count + 1

    try:
        async with httpx.AsyncClient(
            verify=False,
            follow_redirects=True,
            timeout=timeout_ms / 1000.0,
        ) as client:
            for attempt in range(1, attempts_total + 1):
                try:
                    response = await client.get(url)
                    raw_status = response.status_code
                    has_response = True
                    last_status = raw_status if isinstance(raw_status, int) else None
                    last_detail = f"HTTP {raw_status}" if last_status is not None else "无状态码"

                    if track_redirects:
                        redirect_chain = list(response.history or [])
                        last_redirect_count = len(redirect_chain)
                        last_final_url = str(response.url)
                    else:
                        last_redirect_count = 0
                        last_final_url = str(response.url)

                    if extract_title and last_status == 200:
                        content_type = response.headers.get("content-type", "")
                        if "text/html" in content_type.lower():
                            try:
                                soup = BeautifulSoup(response.text, "html.parser")
                                title_tag = soup.find("title")
                                if title_tag:
                                    last_title = title_tag.get_text(strip=True)
                            except Exception:
                                last_title = None

                    if success_checker(last_status, has_response):
                        detail = last_detail if attempt == 1 else f"{last_detail}，重试后成功({attempt}/{attempts_total})"
                        return ProbeResult(
                            thread_name=worker_name,
                            url=url,
                            accessible=True,
                            detail=detail,
                            status_code=last_status,
                            attempts=attempt,
                            title=last_title,
                            redirect_count=last_redirect_count,
                            final_url=last_final_url,
                        )
                except httpx.TimeoutException:
                    last_status = None
                    last_detail = "超时"
                except Exception as exc:
                    last_status = None
                    last_detail = str(exc)
    except Exception as exc:
        last_status = None
        last_detail = f"客户端初始化失败: {exc}"

    detail = f"{last_detail}，已重试 {retry_count} 次" if retry_count > 0 else last_detail
    return ProbeResult(
        thread_name=worker_name,
        url=url,
        accessible=False,
        detail=detail,
        status_code=last_status,
        attempts=attempts_total,
        title=last_title,
        redirect_count=last_redirect_count,
        final_url=last_final_url,
    )


async def run_probe(
    urls: list[str],
    timeout_seconds: int,
    concurrency: int,
    success_checker: Callable[[int | None, bool], bool],
    retry_count: int,
    probe_mode: str = "lightweight",
    extract_title: bool = True,
    track_redirects: bool = True,
    progress_callback: Callable[[int, int, int, int, ProbeResult], None] | None = None,
    is_paused: Callable[[], bool] | None = None,
    is_stopped: Callable[[], bool] | None = None,
    get_concurrency: Callable[[], int] | None = None,
    state_callback: Callable[[str], None] | None = None,
) -> tuple[list[ProbeResult], bool]:
    timeout_ms = timeout_seconds * 1000
    total = len(urls)
    completed = 0
    failed = 0
    results: list[ProbeResult | None] = [None] * total
    stopped = False
    was_paused = False

    def current_concurrency() -> int:
        if get_concurrency is None:
            return concurrency
        try:
            return max(1, int(get_concurrency()))
        except Exception:
            return concurrency

    if probe_mode == "browser":
        return await _run_probe_browser(
            urls=urls,
            timeout_ms=timeout_ms,
            concurrency=concurrency,
            success_checker=success_checker,
            retry_count=retry_count,
            extract_title=extract_title,
            track_redirects=track_redirects,
            progress_callback=progress_callback,
            is_paused=is_paused,
            is_stopped=is_stopped,
            get_concurrency=get_concurrency,
            state_callback=state_callback,
            total=total,
        )
    return await _run_probe_lightweight(
        urls=urls,
        timeout_ms=timeout_ms,
        concurrency=concurrency,
        success_checker=success_checker,
        retry_count=retry_count,
        extract_title=extract_title,
        track_redirects=track_redirects,
        progress_callback=progress_callback,
        is_paused=is_paused,
        is_stopped=is_stopped,
        get_concurrency=get_concurrency,
        state_callback=state_callback,
        total=total,
    )


async def _probe_loop(
    total: int,
    urls: list[str],
    current_concurrency: Callable[[], int],
    is_paused: Callable[[], bool] | None,
    is_stopped: Callable[[], bool] | None,
    state_callback: Callable[[str], None] | None,
    probe_fn: Callable[[int, str, str], "asyncio.Task[tuple[int, ProbeResult]]"],
    progress_callback: Callable[[int, int, int, int, ProbeResult], None] | None,
) -> tuple[list[ProbeResult | None], int, int, bool]:
    results: list[ProbeResult | None] = [None] * total
    completed = 0
    failed = 0
    stopped = False
    was_paused = False
    next_index = 0
    running_tasks: dict[asyncio.Task[tuple[int, ProbeResult]], int] = {}

    while completed < total:
        if is_stopped is not None and is_stopped():
            stopped = True
            break

        paused_now = is_paused is not None and is_paused()
        if paused_now != was_paused and state_callback is not None:
            state_callback("paused" if paused_now else "running")
        was_paused = paused_now

        limit = current_concurrency()
        while not paused_now and next_index < total and len(running_tasks) < limit:
            slot = (next_index % max(1, limit)) + 1
            task = probe_fn(next_index, urls[next_index], f"Worker-{slot:02d}")
            running_tasks[task] = next_index
            next_index += 1

        if not running_tasks:
            await asyncio.sleep(0.1)
            continue

        done, _ = await asyncio.wait(
            running_tasks.keys(),
            timeout=0.1,
            return_when=asyncio.FIRST_COMPLETED,
        )
        for task in done:
            _ = running_tasks.pop(task, None)
            if task.cancelled():
                continue
            index, result = await task
            results[index] = result
            completed += 1
            if not result.accessible:
                failed += 1
            if progress_callback is not None:
                active_workers_display = min(current_concurrency(), max(0, total - completed))
                progress_callback(completed, total, failed, active_workers_display, result)

    if stopped and running_tasks:
        for task in running_tasks:
            _: bool = task.cancel()
        await asyncio.gather(*running_tasks.keys(), return_exceptions=True)

    return results, completed, failed, stopped


async def _run_probe_browser(
    *,
    urls: list[str],
    timeout_ms: int,
    concurrency: int,
    success_checker: Callable[[int | None, bool], bool],
    retry_count: int,
    extract_title: bool,
    track_redirects: bool,
    progress_callback: Callable[[int, int, int, int, ProbeResult], None] | None,
    is_paused: Callable[[], bool] | None,
    is_stopped: Callable[[], bool] | None,
    get_concurrency: Callable[[], int] | None,
    state_callback: Callable[[str], None] | None,
    total: int,
) -> tuple[list[ProbeResult], bool]:
    def current_concurrency() -> int:
        if get_concurrency is None:
            return concurrency
        try:
            return max(1, int(get_concurrency()))
        except Exception:
            return concurrency

    async with async_playwright() as p:
        browser = await p.chromium.launch(headless=True)
        context = await browser.new_context(ignore_https_errors=True)
        try:
            async def wrapped_probe(index: int, url: str, worker_name: str):
                result = await probe_url(context, worker_name, url, timeout_ms, success_checker, retry_count)
                if extract_title and result.accessible:
                    try:
                        page = await context.new_page()
                        await page.goto(url, wait_until="domcontentloaded", timeout=timeout_ms)
                        result.title = await page.title()
                        await page.close()
                    except Exception:
                        pass
                return index, result

            def probe_fn(index: int, url: str, worker_name: str):
                return asyncio.create_task(
                    wrapped_probe(index, url, worker_name),
                    name=worker_name,
                )

            results, completed, failed, stopped = await _probe_loop(
                total=total,
                urls=urls,
                current_concurrency=current_concurrency,
                is_paused=is_paused,
                is_stopped=is_stopped,
                state_callback=state_callback,
                probe_fn=probe_fn,
                progress_callback=progress_callback,
            )
            return [item for item in results if item is not None], stopped
        finally:
            await context.close()
            await browser.close()


async def _run_probe_lightweight(
    *,
    urls: list[str],
    timeout_ms: int,
    concurrency: int,
    success_checker: Callable[[int | None, bool], bool],
    retry_count: int,
    extract_title: bool,
    track_redirects: bool,
    progress_callback: Callable[[int, int, int, int, ProbeResult], None] | None,
    is_paused: Callable[[], bool] | None,
    is_stopped: Callable[[], bool] | None,
    get_concurrency: Callable[[], int] | None,
    state_callback: Callable[[str], None] | None,
    total: int,
) -> tuple[list[ProbeResult], bool]:
    def current_concurrency() -> int:
        if get_concurrency is None:
            return concurrency
        try:
            return max(1, int(get_concurrency()))
        except Exception:
            return concurrency

    async def wrapped_probe(index: int, url: str, worker_name: str):
        result = await probe_url_httpx(
            worker_name, url, timeout_ms, success_checker, retry_count, extract_title, track_redirects
        )
        return index, result

    def probe_fn(index: int, url: str, worker_name: str):
        return asyncio.create_task(
            wrapped_probe(index, url, worker_name),
            name=worker_name,
        )

    results, completed, failed, stopped = await _probe_loop(
        total=total,
        urls=urls,
        current_concurrency=current_concurrency,
        is_paused=is_paused,
        is_stopped=is_stopped,
        state_callback=state_callback,
        probe_fn=probe_fn,
        progress_callback=progress_callback,
    )
    return [item for item in results if item is not None], stopped


def prepare_column_options(sheet: Worksheet, has_header: bool) -> list[tuple[str, int]]:
    options: list[tuple[str, int]] = []
    for col_idx in range(1, sheet.max_column + 1):
        header_value = sheet.cell(row=1, column=col_idx).value if has_header else None
        display = str(header_value).strip() if header_value is not None else ""
        if not display:
            display = f"第 {col_idx} 列"
        options.append((display, col_idx))
    return options


def write_results_and_stats(
    wb: Workbook,
    selected_sheet_name: str,
    has_header: bool,
    probe_results: list[ProbeResult],
) -> tuple[int, int]:
    sheet: Worksheet = wb[selected_sheet_name]
    result_col = sheet.max_column + 1
    detail_col = result_col + 1
    title_col = detail_col + 1
    redirect_col = title_col + 1
    final_url_col = redirect_col + 1
    start_row = 2 if has_header else 1
    if has_header:
        sheet.cell(row=1, column=result_col, value="是否可访问")
        sheet.cell(row=1, column=detail_col, value="访问详情")
        sheet.cell(row=1, column=title_col, value="网站标题")
        sheet.cell(row=1, column=redirect_col, value="重定向次数")
        sheet.cell(row=1, column=final_url_col, value="最终URL")
    success_count = 0
    total_count = len(probe_results)
    for idx, item in enumerate(probe_results):
        row = start_row + idx
        sheet.cell(row=row, column=result_col, value="是" if item.accessible else "否")
        sheet.cell(row=row, column=detail_col, value=item.detail)
        sheet.cell(row=row, column=title_col, value=item.title or "")
        sheet.cell(row=row, column=redirect_col, value=item.redirect_count)
        sheet.cell(row=row, column=final_url_col, value=item.final_url or "")
        if item.accessible:
            success_count += 1
    stats_name = "拨测统计"
    if stats_name in wb.sheetnames:
        del wb[stats_name]
    stat_sheet: Worksheet = wb.create_sheet(stats_name)
    ratio = (success_count / total_count) if total_count else 0
    stat_sheet["A1"] = "统计项"
    stat_sheet["B1"] = "值"
    stat_sheet["A2"] = "总拨测数"
    stat_sheet["B2"] = total_count
    stat_sheet["A3"] = "可访问数"
    stat_sheet["B3"] = success_count
    stat_sheet["A4"] = "不可访问数"
    stat_sheet["B4"] = total_count - success_count
    stat_sheet["A5"] = "可访问比例"  # type: ignore
    stat_sheet["B5"] = f"{ratio:.2%}"  # type: ignore
    stat_sheet["A7"] = "生成时间"  # type: ignore
    stat_sheet["B7"] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")  # type: ignore
    return success_count, total_count


def execute_probe_job(
    config: ProbeConfig,
    *,
    progress_callback: Callable[[int, int, int, int, ProbeResult], None] | None = None,
    is_paused: Callable[[], bool] | None = None,
    is_stopped: Callable[[], bool] | None = None,
    get_concurrency: Callable[[], int] | None = None,
    state_callback: Callable[[str], None] | None = None,
) -> ExecuteProbeResult:
    wb, _ = load_workbook_from_input(config.file_path, config.password, config.use_office_fallback)
    if config.sheet_name not in wb.sheetnames:
        raise RuntimeError(f"Sheet 不存在: {config.sheet_name!r}，可用: {', '.join(wb.sheetnames)}")
    sheet: Worksheet = wb[config.sheet_name]
    urls = extract_urls_from_sheet(sheet, config.url_col_idx, config.has_header)
    if not urls:
        raise RuntimeError("目标列没有可拨测数据。")

    success_checker = build_success_checker(config.success_mode, config.custom_codes_input)
    results, stopped = asyncio.run(
        run_probe(
            urls=urls,
            timeout_seconds=config.timeout_seconds,
            concurrency=config.concurrency,
            success_checker=success_checker,
            retry_count=config.retry_count,
            probe_mode=config.probe_mode,
            extract_title=config.extract_title,
            track_redirects=config.track_redirects,
            progress_callback=progress_callback,
            is_paused=is_paused,
            is_stopped=is_stopped,
            get_concurrency=get_concurrency,
            state_callback=state_callback,
        )
    )
    failed_count = sum(not item.accessible for item in results)
    if stopped:
        if results:
            write_results_and_stats(wb, config.sheet_name, config.has_header, results)
            wb.save(config.output_path)
        return ExecuteProbeResult(
            outcome="aborted",
            success_count=sum(item.accessible for item in results),
            total_count=len(urls),
            failed_count=failed_count,
            output_path=config.output_path,
        )

    success_count, total_count = write_results_and_stats(wb, config.sheet_name, config.has_header, results)
    wb.save(config.output_path)
    return ExecuteProbeResult(
        outcome="success",
        success_count=success_count,
        total_count=total_count,
        failed_count=total_count - success_count,
        output_path=config.output_path,
    )
