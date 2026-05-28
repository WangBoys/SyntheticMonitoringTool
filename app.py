from __future__ import annotations

import os
import re
import sys
import threading
import webbrowser
from datetime import datetime
from time import monotonic
from typing import final

try:
    from typing import override  # Python 3.12+
except ImportError:
    from typing_extensions import override

from openpyxl import Workbook
from openpyxl.worksheet.worksheet import Worksheet
from PySide6 import QtCore, QtGui, QtWidgets

from probe_core import (
    DEFAULT_CONCURRENCY,
    DEFAULT_RETRY_COUNT,
    DEFAULT_TIMEOUT_SECONDS,
    ExecuteProbeResult,
    ProbeConfig,
    ProbeResult,
    detect_public_ip_and_isp,
    execute_probe_job,
    get_app_base_dir,
    is_valid_url_or_domain,
    load_workbook_from_input,
    normalize_url,
    prepare_column_options,
)


MAX_LIVE_LOG_ROWS = 2000


def get_app_icon_path() -> str:
    base_dir = get_app_base_dir()
    candidates = [
        os.path.join(base_dir, "app_icon.ico"),
        os.path.join(os.path.dirname(os.path.abspath(__file__)), "app_icon.ico"),
    ]
    meipass = getattr(sys, "_MEIPASS", "")
    if meipass:
        candidates.append(os.path.join(meipass, "app_icon.ico"))
    for path in candidates:
        if os.path.isfile(path):
            return path
    return ""


class FilePathEdit(QtWidgets.QLineEdit):
    def __init__(self, parent: QtWidgets.QWidget | None = None):
        super().__init__(parent)
        self.setAcceptDrops(True)

    @override
    def dragEnterEvent(self, event: QtGui.QDragEnterEvent) -> None:  # type: ignore[override]
        if event.mimeData().hasUrls():
            event.acceptProposedAction()
        else:
            super().dragEnterEvent(event)

    @override
    def dropEvent(self, event: QtGui.QDropEvent) -> None:  # type: ignore[override]
        urls = event.mimeData().urls()
        if urls:
            local_path = urls[0].toLocalFile()
            if local_path:
                self.setText(local_path)
        event.acceptProposedAction()


class ProbeWorker(QtCore.QThread):
    progress: final = QtCore.Signal(int, int, int, int, object)  # type: ignore[valid-type]
    success: final = QtCore.Signal(int, int, str)  # type: ignore[valid-type]
    aborted: final = QtCore.Signal(int, int, int)  # type: ignore[valid-type]
    failed: final = QtCore.Signal(str)  # type: ignore[valid-type]
    state_changed: final = QtCore.Signal(str)  # type: ignore[valid-type]

    def __init__(self, config: ProbeConfig):
        super().__init__()
        self.config: final = config  # type: ignore[valid-type]
        self._pause_event: final = threading.Event()  # type: ignore[valid-type]
        self._stop_event: final = threading.Event()  # type: ignore[valid-type]
        self._concurrency_lock: final = threading.Lock()  # type: ignore[valid-type]
        self._current_concurrency: final = config.concurrency  # type: ignore[valid-type]

    def pause_probe(self) -> None:
        self._pause_event.set()

    def resume_probe(self) -> None:
        self._pause_event.clear()

    def stop_probe(self) -> None:
        self._stop_event.set()
        self._pause_event.clear()

    def update_concurrency(self, concurrency: int) -> None:
        with self._concurrency_lock:
            self._current_concurrency = max(1, int(concurrency))

    def _get_concurrency(self) -> int:
        with self._concurrency_lock:
            return self._current_concurrency

    @override
    def run(self) -> None:
        try:
            result: ExecuteProbeResult = execute_probe_job(
                self.config,
                progress_callback=lambda c, t, f, a, r: self.progress.emit(c, t, f, a, r),
                is_paused=self._pause_event.is_set,
                is_stopped=self._stop_event.is_set,
                get_concurrency=self._get_concurrency,
                state_callback=lambda state: self.state_changed.emit(state),
            )
            if result.outcome == "aborted":
                self.aborted.emit(result.success_count + result.failed_count, result.total_count, result.failed_count)
                return
            self.success.emit(result.success_count, result.total_count, result.output_path)
        except Exception as exc:
            self.failed.emit(str(exc))


class IpDetectWorker(QtCore.QThread):
    success: final = QtCore.Signal(str, str, str)  # type: ignore[valid-type]
    failed: final = QtCore.Signal(str)  # type: ignore[valid-type]

    @override
    def run(self) -> None:  # type: ignore[override]
        try:
            ip, isp, location = detect_public_ip_and_isp()
            self.success.emit(ip, isp, location)
        except Exception as exc:
            self.failed.emit(str(exc))


class MainWindow(QtWidgets.QMainWindow):
    def __init__(self):
        super().__init__()
        self.setWindowTitle("网址拨测可视化工具 (PySide6)")
        self.resize(1000, 760)
        self.workbook: final = None  # type: ignore[valid-type]
        self._loaded_file_path: final = ""  # type: ignore[valid-type]
        self._file_password: final = ""  # type: ignore[valid-type]
        self.worker: ProbeWorker | None = None
        self.ip_worker: IpDetectWorker | None = None
        self._run_state: final = "idle"  # type: ignore[valid-type]
        self._url_column_valid: final = False  # type: ignore[valid-type]
        self._run_start_dt: datetime | None = None
        self._run_start_monotonic: float | None = None
        self._log_entries: list[ProbeResult] = []
        self._build_ui()
        self._apply_ui_state("idle")
        self._refresh_public_ip()

    def _build_ui(self) -> None:
        central = QtWidgets.QWidget()
        self.setCentralWidget(central)
        root = QtWidgets.QVBoxLayout(central)

        file_box = QtWidgets.QGroupBox("文件输入（支持输入/拖拽/弹窗）")
        file_layout = QtWidgets.QGridLayout(file_box)
        self.file_edit: final = FilePathEdit()  # type: ignore[valid-type]
        self.file_edit.setPlaceholderText(r"E:\path\to\target.xlsx（可直接拖拽文件到这里）")
        _: QtCore.QMetaObject.Connection = self.file_edit.textChanged.connect(self._on_file_path_changed)
        self.btn_browse_file: final = QtWidgets.QPushButton("弹窗选择文件")  # type: ignore[valid-type]
        _: QtCore.QMetaObject.Connection = self.btn_browse_file.clicked.connect(self._on_browse_file)
        self.btn_browse_file.setSizePolicy(QtWidgets.QSizePolicy.Policy.Expanding, QtWidgets.QSizePolicy.Policy.Fixed)
        file_layout.addWidget(QtWidgets.QLabel("Excel 文件路径"), 0, 0)
        file_layout.addWidget(self.file_edit, 0, 1, 1, 2)
        file_layout.addWidget(self.btn_browse_file, 0, 3)
        file_layout.setColumnStretch(1, 8)
        file_layout.setColumnStretch(2, 2)
        file_layout.setColumnStretch(3, 3)

        self.office_fallback: final = QtWidgets.QCheckBox("密码解密失败时尝试 Office/WPS")  # type: ignore[valid-type]
        self.file_status: final = QtWidgets.QLabel("文件结构: 未读取")  # type: ignore[valid-type]
        file_layout.addWidget(self.office_fallback, 1, 1, 1, 2)
        file_layout.addWidget(self.file_status, 2, 0, 1, 3)
        self.public_ip_label: final = QtWidgets.QLabel("公网 IP: 检测中...")  # type: ignore[valid-type]
        self.public_isp_label: final = QtWidgets.QLabel("运营商: 检测中...")  # type: ignore[valid-type]
        self.public_location_label: final = QtWidgets.QLabel("国家/省份/城市: 检测中...")  # type: ignore[valid-type]
        self.btn_refresh_ip: final = QtWidgets.QPushButton("刷新 IP/运营商/地区")  # type: ignore[valid-type]
        _: QtCore.QMetaObject.Connection = self.btn_refresh_ip.clicked.connect(self._refresh_public_ip)
        file_layout.addWidget(self.public_ip_label, 3, 0)
        file_layout.addWidget(self.public_isp_label, 3, 1)
        file_layout.addWidget(self.public_location_label, 3, 2)
        file_layout.addWidget(self.btn_refresh_ip, 3, 3)
        root.addWidget(file_box)

        config_box = QtWidgets.QGroupBox("拨测配置")
        config_layout = QtWidgets.QGridLayout(config_box)
        self.sheet_combo: final = QtWidgets.QComboBox()  # type: ignore[valid-type]
        _: QtCore.QMetaObject.Connection = self.sheet_combo.currentIndexChanged.connect(self._refresh_column_options)
        self.has_header: final = QtWidgets.QCheckBox("第一行是表头")  # type: ignore[valid-type]
        self.has_header.setChecked(True)
        _: QtCore.QMetaObject.Connection = self.has_header.stateChanged.connect(self._refresh_column_options)
        self.column_combo: final = QtWidgets.QComboBox()  # type: ignore[valid-type]
        _: QtCore.QMetaObject.Connection = self.column_combo.currentIndexChanged.connect(self._validate_selected_url_column)
        config_layout.addWidget(QtWidgets.QLabel("Sheet"), 0, 0)
        config_layout.addWidget(self.sheet_combo, 0, 1)
        config_layout.addWidget(self.has_header, 0, 2)
        config_layout.addWidget(QtWidgets.QLabel("网址列"), 1, 0)
        config_layout.addWidget(self.column_combo, 1, 1, 1, 2)
        self.url_column_status_label: final = QtWidgets.QLabel("网址列校验: 未校验")  # type: ignore[valid-type]
        config_layout.addWidget(self.url_column_status_label, 1, 3, 1, 3)

        self.concurrency_spin: final = QtWidgets.QSpinBox()  # type: ignore[valid-type]
        self.concurrency_spin.setRange(1, 100)
        self.concurrency_spin.setValue(DEFAULT_CONCURRENCY)
        _: QtCore.QMetaObject.Connection = self.concurrency_spin.valueChanged.connect(self._update_thread_hint)
        _: QtCore.QMetaObject.Connection = self.concurrency_spin.valueChanged.connect(self._on_concurrency_changed)
        self.timeout_spin: final = QtWidgets.QSpinBox()  # type: ignore[valid-type]
        self.timeout_spin.setRange(1, 300)
        self.timeout_spin.setValue(DEFAULT_TIMEOUT_SECONDS)
        self.retry_spin: final = QtWidgets.QSpinBox()  # type: ignore[valid-type]
        self.retry_spin.setRange(0, 5)
        self.retry_spin.setValue(DEFAULT_RETRY_COUNT)
        config_layout.addWidget(QtWidgets.QLabel("并发数"), 2, 0)
        config_layout.addWidget(self.concurrency_spin, 2, 1)
        config_layout.addWidget(QtWidgets.QLabel("超时(秒)"), 2, 2)
        config_layout.addWidget(self.timeout_spin, 2, 3)
        config_layout.addWidget(QtWidgets.QLabel("失败重试"), 2, 4)
        config_layout.addWidget(self.retry_spin, 2, 5)
        self.thread_hint: final = QtWidgets.QLabel()
        config_layout.addWidget(self.thread_hint, 2, 6)
        self._update_thread_hint()

        self.success_mode_combo: final = QtWidgets.QComboBox()  # pyright: ignore[reportUnannotatedClassAttribute]
        self.success_mode_combo.addItems(["仅 2xx/3xx", "2xx/3xx + 无状态码也算成功", "2xx/3xx/4xx/5xx（任意状态码）", "自定义 HTTP 状态码"])
        self.success_mode_combo.setCurrentIndex(2)
        _: QtCore.QMetaObject.Connection = self.success_mode_combo.currentIndexChanged.connect(self._toggle_custom_codes)
        self.custom_codes_edit: final = QtWidgets.QLineEdit("2xx,3xx,200,204")  # pyright: ignore[reportUnannotatedClassAttribute]
        self.custom_codes_edit.setToolTip("支持精确码和段规则混填，例如: 200,204,3xx,5xx")
        self.custom_codes_edit.setEnabled(False)
        config_layout.addWidget(QtWidgets.QLabel("成功规则"), 3, 0)
        config_layout.addWidget(self.success_mode_combo, 3, 1, 1, 2)
        config_layout.addWidget(QtWidgets.QLabel("自定义状态码"), 3, 3)
        config_layout.addWidget(self.custom_codes_edit, 3, 4, 1, 2)

        self.probe_mode_combo: final = QtWidgets.QComboBox()  # type: ignore[valid-type]
        self.probe_mode_combo.addItems(["轻量模式", "浏览器模式"])
        self.probe_mode_combo.setCurrentIndex(0)
        self.probe_mode_combo.setToolTip("轻量模式：使用 httpx，速度快；浏览器模式：使用 Playwright，支持 JS 渲染")
        config_layout.addWidget(QtWidgets.QLabel("探测模式"), 4, 0)
        config_layout.addWidget(self.probe_mode_combo, 4, 1, 1, 2)

        self.extract_title_check: final = QtWidgets.QCheckBox("提取网站标题")  # type: ignore[valid-type]
        self.extract_title_check.setChecked(True)
        self.extract_title_check.setToolTip("从响应中解析并提取网页标题")
        config_layout.addWidget(self.extract_title_check, 4, 3, 1, 2)

        self.track_redirects_check: final = QtWidgets.QCheckBox("跟踪重定向")  # type: ignore[valid-type]
        self.track_redirects_check.setChecked(True)
        self.track_redirects_check.setToolTip("记录重定向次数和最终 URL")
        config_layout.addWidget(self.track_redirects_check, 4, 5, 1, 2)

        self.output_edit: final = QtWidgets.QLineEdit()  # type: ignore[valid-type]
        _: QtCore.QMetaObject.Connection = self.output_edit.textChanged.connect(self._update_open_output_button_state)
        self.btn_output: final = QtWidgets.QPushButton("选择输出路径")  # type: ignore[valid-type]
        _: QtCore.QMetaObject.Connection = self.btn_output.clicked.connect(self._on_browse_output)
        self.btn_open_output: final = QtWidgets.QPushButton("打开输出文件")  # type: ignore[valid-type]
        _: QtCore.QMetaObject.Connection = self.btn_open_output.clicked.connect(self._on_open_output)
        config_layout.addWidget(QtWidgets.QLabel("输出文件"), 4, 0)
        config_layout.addWidget(self.output_edit, 4, 1, 1, 4)
        config_layout.addWidget(self.btn_output, 4, 5)
        config_layout.addWidget(self.btn_open_output, 4, 6)
        self._update_open_output_button_state()

        root.addWidget(config_box)

        action_layout = QtWidgets.QHBoxLayout()
        self.btn_start: final = QtWidgets.QPushButton("开始拨测")  # type: ignore[valid-type]
        _: QtCore.QMetaObject.Connection = self.btn_start.clicked.connect(self._on_start)
        self.btn_pause: final = QtWidgets.QPushButton("暂停拨测")  # type: ignore[valid-type]
        _: QtCore.QMetaObject.Connection = self.btn_pause.clicked.connect(self._on_pause_resume)
        self.btn_stop: final = QtWidgets.QPushButton("中止拨测")  # type: ignore[valid-type]
        _: QtCore.QMetaObject.Connection = self.btn_stop.clicked.connect(self._on_stop)
        self.progress: final = QtWidgets.QProgressBar()  # type: ignore[valid-type]
        self.progress.setRange(0, 100)
        self.progress_text: final = QtWidgets.QLabel("进度: 0/0，失败: 0")  # type: ignore[valid-type]
        self.metric_label: final = QtWidgets.QLabel("可访问比例: -")  # type: ignore[valid-type]
        action_layout.addWidget(self.btn_start)
        action_layout.addWidget(self.btn_pause)
        action_layout.addWidget(self.btn_stop)
        action_layout.addWidget(self.progress)
        action_layout.addWidget(self.progress_text)
        action_layout.addWidget(self.metric_label)
        root.addLayout(action_layout)

        time_layout = QtWidgets.QHBoxLayout()
        self.start_time_label: final = QtWidgets.QLabel("开始时间: -")  # type: ignore[valid-type]
        self.elapsed_time_label: final = QtWidgets.QLabel("实时耗时: 00:00:00")  # type: ignore[valid-type]
        self.end_time_label: final = QtWidgets.QLabel("结束时间: -")  # type: ignore[valid-type]
        time_layout.addWidget(self.start_time_label)
        time_layout.addWidget(self.elapsed_time_label)
        time_layout.addWidget(self.end_time_label)
        time_layout.addStretch(1)
        root.addLayout(time_layout)

        filter_layout = QtWidgets.QHBoxLayout()
        self.status_filter_edit: final = QtWidgets.QLineEdit()
        self.status_filter_edit.setPlaceholderText("状态码筛选: 200,3xx,无状态码")
        _: QtCore.QMetaObject.Connection = self.status_filter_edit.textChanged.connect(self._refresh_log_table)
        self.access_filter_combo: final = QtWidgets.QComboBox()  # type: ignore[valid-type]
        self.access_filter_combo.addItems(["是否可访问: 全部", "是否可访问: 是", "是否可访问: 否"])
        _: QtCore.QMetaObject.Connection = self.access_filter_combo.currentIndexChanged.connect(self._refresh_log_table)
        self.detail_filter_edit: final = QtWidgets.QLineEdit()  # type: ignore[valid-type]
        self.detail_filter_edit.setPlaceholderText("详情筛选关键字")
        _: QtCore.QMetaObject.Connection = self.detail_filter_edit.textChanged.connect(self._refresh_log_table)
        filter_layout.addWidget(self.status_filter_edit)
        filter_layout.addWidget(self.access_filter_combo)
        filter_layout.addWidget(self.detail_filter_edit)
        root.addLayout(filter_layout)

        self.elapsed_timer: final = QtCore.QTimer(self)  # type: ignore[valid-type]
        self.elapsed_timer.setInterval(1000)
        _: QtCore.QMetaObject.Connection = self.elapsed_timer.timeout.connect(self._update_elapsed_time)

        self.preview_table: final = QtWidgets.QTableWidget(0, 8)  # type: ignore[valid-type]
        self.preview_table.setHorizontalHeaderLabels(["线程", "URL", "状态码", "是否可访问", "详情", "网站标题", "重定向次数", "最终URL"])
        self.preview_table.horizontalHeader().setStretchLastSection(True)
        self.preview_table.setEditTriggers(QtWidgets.QAbstractItemView.EditTrigger.NoEditTriggers)
        self.preview_table.setContextMenuPolicy(QtCore.Qt.ContextMenuPolicy.CustomContextMenu)
        _: object = self.preview_table.customContextMenuRequested.connect(self._show_log_context_menu)
        root.addWidget(self.preview_table)

    def _on_file_path_changed(self) -> None:
        path = self.file_edit.text().strip()
        if path and path != self._loaded_file_path:
            self.workbook = None
            self._url_column_valid = False
            self._file_password = ""
            self.sheet_combo.clear()
            self.column_combo.clear()
            self.file_status.setText("文件结构: 未读取")
        if path:
            if not self.output_edit.text().strip():
                self.output_edit.setText(os.path.splitext(path)[0] + "_monitoring_result.xlsx")
            if os.path.isfile(path):
                self._auto_load_structure(path)
        self._apply_ui_state(self._run_state)

    def _refresh_public_ip(self) -> None:
        if self.ip_worker is not None:
            return
        self.public_ip_label.setText("公网 IP: 检测中...")
        self.public_isp_label.setText("运营商: 检测中...")
        self.public_location_label.setText("国家/省份/城市: 检测中...")
        self.btn_refresh_ip.setEnabled(False)
        self.ip_worker = IpDetectWorker()
        _: object = self.ip_worker.success.connect(self._on_ip_detect_success)
        _: object = self.ip_worker.failed.connect(self._on_ip_detect_failed)
        _: object = self.ip_worker.finished.connect(self._on_ip_detect_finished)
        self.ip_worker.start()

    def _on_ip_detect_success(self, ip: str, isp: str, location: str) -> None:
        self.public_ip_label.setText(f"公网 IP: {ip}")
        self.public_isp_label.setText(f"运营商: {isp}")
        self.public_location_label.setText(f"国家/省份/城市: {location}")

    def _on_ip_detect_failed(self, message: str) -> None:
        self.public_ip_label.setText("公网 IP: 检测失败")
        self.public_isp_label.setText("运营商: 检测失败（网络受限）")
        self.public_location_label.setText("国家/省份/城市: 检测失败")

    def _on_ip_detect_finished(self) -> None:
        self.ip_worker = None
        self.btn_refresh_ip.setEnabled(True)

    def _on_browse_file(self) -> None:
        path, _ = QtWidgets.QFileDialog.getOpenFileName(
            self,
            "选择 Excel 文件",
            "",
            "Excel Files (*.xlsx *.xlsm *.xltx *.xltm);;All Files (*.*)",
        )
        if path:
            self.file_edit.setText(path)
            self._auto_load_structure(path)

    def _on_browse_output(self) -> None:
        default_name = self.output_edit.text().strip() or "monitoring_result.xlsx"
        path, _ = QtWidgets.QFileDialog.getSaveFileName(self, "选择输出路径", default_name, "Excel Files (*.xlsx)")
        if path:
            self.output_edit.setText(path)

    def _on_open_output(self) -> None:
        path = self.output_edit.text().strip()
        if not path:
            _: object = QtWidgets.QMessageBox.warning(self, "提示", "请先设置输出文件路径。")
            return
        if not os.path.isfile(path):
            _: object = QtWidgets.QMessageBox.warning(self, "提示", f"输出文件不存在:\n{path}")
            return
        opened = QtGui.QDesktopServices.openUrl(QtCore.QUrl.fromLocalFile(path))
        if not opened:
            _ = QtWidgets.QMessageBox.warning(self, "提示", "打开输出文件失败，请检查系统默认关联设置。")

    def _update_open_output_button_state(self) -> None:
        path = self.output_edit.text().strip()
        self.btn_open_output.setEnabled(bool(path) and os.path.isfile(path))

    def _status_code_text(self, item: ProbeResult) -> str:
        return str(item.status_code) if item.status_code is not None else "无状态码"

    def _matches_status_filter(self, item: ProbeResult) -> bool:
        raw = self.status_filter_edit.text().strip()
        if not raw:
            return True
        status_text = self._status_code_text(item).lower()
        tokens = [x.strip().lower() for x in raw.replace("，", ",").split(",") if x.strip()]
        if not tokens:
            return True
        for token in tokens:
            if token in {"无状态码", "none", "null", "-"} and item.status_code is None:
                return True
            if re.fullmatch(r"[1-5]xx", token):
                if item.status_code is not None and item.status_code // 100 == int(token[0]):
                    return True
                continue
            if re.fullmatch(r"\d{3}", token):
                if item.status_code == int(token):
                    return True
                continue
            if token in status_text:
                return True
        return False

    def _matches_filters(self, item: ProbeResult) -> bool:
        if not self._matches_status_filter(item):
            return False
        access_mode = self.access_filter_combo.currentIndex()
        if access_mode == 1 and not item.accessible:
            return False
        if access_mode == 2 and item.accessible:
            return False
        detail_kw = self.detail_filter_edit.text().strip().lower()
        if detail_kw and detail_kw not in item.detail.lower():
            return False
        return True

    def _refresh_log_table(self) -> None:
        filtered = [entry for entry in self._log_entries if self._matches_filters(entry)]
        self.preview_table.setRowCount(len(filtered))
        for row, item in enumerate(filtered):
            self.preview_table.setItem(row, 0, QtWidgets.QTableWidgetItem(item.thread_name))
            url_item = QtWidgets.QTableWidgetItem(item.url)
            url_item.setToolTip(item.url)
            self.preview_table.setItem(row, 1, url_item)
            self.preview_table.setItem(row, 2, QtWidgets.QTableWidgetItem(self._status_code_text(item)))
            self.preview_table.setItem(row, 3, QtWidgets.QTableWidgetItem("是" if item.accessible else "否"))
            self.preview_table.setItem(row, 4, QtWidgets.QTableWidgetItem(item.detail))
            self.preview_table.setItem(row, 5, QtWidgets.QTableWidgetItem(item.title or ""))
            self.preview_table.setItem(row, 6, QtWidgets.QTableWidgetItem(str(item.redirect_count)))
            self.preview_table.setItem(row, 7, QtWidgets.QTableWidgetItem(item.final_url or ""))
        if filtered:
            self.preview_table.scrollToBottom()

    def _format_elapsed(self, elapsed_seconds: int) -> str:
        hours, rem = divmod(max(0, int(elapsed_seconds)), 3600)
        minutes, seconds = divmod(rem, 60)
        return f"{hours:02d}:{minutes:02d}:{seconds:02d}"

    def _mark_run_started(self) -> None:
        self._run_start_dt = datetime.now()
        self._run_start_monotonic = monotonic()
        self.start_time_label.setText(f"开始时间: {self._run_start_dt.strftime('%Y-%m-%d %H:%M:%S')}")
        self.elapsed_time_label.setText("实时耗时: 00:00:00")
        self.end_time_label.setText("结束时间: -")
        self.elapsed_timer.start()

    def _mark_run_finished(self) -> None:
        self.elapsed_timer.stop()
        self._update_elapsed_time()
        self.end_time_label.setText(f"结束时间: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")

    def _update_elapsed_time(self) -> None:
        if self._run_start_monotonic is None:
            self.elapsed_time_label.setText("实时耗时: 00:00:00")
            return
        elapsed = int(monotonic() - self._run_start_monotonic)
        self.elapsed_time_label.setText(f"实时耗时: {self._format_elapsed(elapsed)}")

    def _auto_load_structure(self, file_path: str) -> None:
        path = file_path.strip()
        if not path or not os.path.isfile(path):
            return
        if path == self._loaded_file_path and self.workbook is not None:
            return
        wb: Workbook | None = None
        try:
            wb, _ = load_workbook_from_input(path, "", self.office_fallback.isChecked())
        except Exception as first_error:
            ask = QtWidgets.QMessageBox.question(
                self,
                "读取失败",
                f"读取文件失败，可能是加密文件。\n是否输入密码重试？\n\n错误: {first_error}",
                QtWidgets.QMessageBox.StandardButton.Yes | QtWidgets.QMessageBox.StandardButton.No,
                QtWidgets.QMessageBox.StandardButton.Yes,
            )
            if ask != QtWidgets.QMessageBox.StandardButton.Yes:
                self.file_status.setText("文件结构: 读取失败")
                return
            while True:
                password, ok = QtWidgets.QInputDialog.getText(
                    self,
                    "输入文件密码",
                    "请输入该 Excel 的密码：",
                    QtWidgets.QLineEdit.EchoMode.Password,
                )
                if not ok:  # type: ignore
                    self.file_status.setText("文件结构: 已取消密码输入")
                    return
                if not password:
                    _: object = QtWidgets.QMessageBox.warning(self, "提示", "密码不能为空。")
                    continue
                try:
                    wb, _ = load_workbook_from_input(path, password, self.office_fallback.isChecked())
                    self._file_password = password
                    break
                except Exception as pwd_error:
                    retry = QtWidgets.QMessageBox.question(  # type: ignore
                        self,
                        "密码错误或解密失败",
                        f"解密失败: {pwd_error}\n是否重新输入密码？",
                        QtWidgets.QMessageBox.StandardButton.Yes | QtWidgets.QMessageBox.StandardButton.No,
                        QtWidgets.QMessageBox.StandardButton.Yes,
                    )
                    if retry != QtWidgets.QMessageBox.StandardButton.Yes:  # type: ignore[comparison-overlap]
                        self.file_status.setText("文件结构: 读取失败")
                        return

        if wb is None:
            self.file_status.setText("文件结构: 读取失败")
            return

        self.workbook = wb
        self._loaded_file_path = path
        self.sheet_combo.clear()
        self.sheet_combo.addItems(wb.sheetnames)
        self._refresh_column_options()
        self.file_status.setText(f"文件结构: 已读取 {len(wb.sheetnames)} 个 Sheet")
        self._apply_ui_state(self._run_state)

    def _refresh_column_options(self) -> None:
        if self.workbook is None or self.sheet_combo.count() == 0:
            self._url_column_valid = False
            self.url_column_status_label.setText("网址列校验: 未校验")
            return
        sheet_name = self.sheet_combo.currentText()
        sheet: Worksheet = self.workbook[sheet_name]
        options = prepare_column_options(sheet, self.has_header.isChecked())
        self.column_combo.clear()
        for label, idx in options:
            self.column_combo.addItem(label, idx)
        self._validate_selected_url_column()

    def _validate_selected_url_column(self) -> None:
        if self.workbook is None or self.sheet_combo.count() == 0 or self.column_combo.currentData() is None:
            self._url_column_valid = False
            self.url_column_status_label.setText("网址列校验: 未校验")
            self._apply_ui_state(self._run_state)
            return

        sheet_name = self.sheet_combo.currentText()
        sheet: Worksheet = self.workbook[sheet_name]
        col_idx = int(self.column_combo.currentData())
        start_row = 2 if self.has_header.isChecked() else 1
        max_check_rows = 500
        checked = 0
        found_valid = False

        for row_idx in range(start_row, sheet.max_row + 1):
            raw = sheet.cell(row=row_idx, column=col_idx).value
            checked += 1
            if is_valid_url_or_domain(raw):
                found_valid = True
                break
            if checked >= max_check_rows:
                break

        self._url_column_valid = found_valid
        if found_valid:
            self.url_column_status_label.setText(f"网址列校验: 通过（已检测前 {checked} 行）")
        else:
            self.url_column_status_label.setText("网址列校验: 需重新选择包含网址/域名的列")
        self._apply_ui_state(self._run_state)

    def _toggle_custom_codes(self) -> None:
        is_custom = self.success_mode_combo.currentText() == "自定义 HTTP 状态码"
        self.custom_codes_edit.setEnabled(is_custom and self._run_state == "idle")

    def _apply_ui_state(self, state: str) -> None:
        self._run_state = state
        has_structure = self.workbook is not None and self.column_combo.count() > 0
        can_start_probe = has_structure and self._url_column_valid
        allow_all = state == "idle"
        allow_concurrency = state in {"idle", "paused"}

        self.file_edit.setEnabled(allow_all)
        self.btn_browse_file.setEnabled(allow_all)
        self.office_fallback.setEnabled(allow_all)
        self.sheet_combo.setEnabled(allow_all)
        self.has_header.setEnabled(allow_all)
        self.column_combo.setEnabled(allow_all)
        self.timeout_spin.setEnabled(allow_all)
        self.retry_spin.setEnabled(allow_all)
        self.success_mode_combo.setEnabled(allow_all)
        self.probe_mode_combo.setEnabled(allow_all)
        self.extract_title_check.setEnabled(allow_all)
        self.track_redirects_check.setEnabled(allow_all)
        self.output_edit.setEnabled(allow_all)
        self.btn_output.setEnabled(allow_all)
        self._update_open_output_button_state()
        self.concurrency_spin.setEnabled(allow_concurrency)
        self.custom_codes_edit.setEnabled(
            allow_all and self.success_mode_combo.currentText() == "自定义 HTTP 状态码"
        )

        self.btn_start.setEnabled(allow_all and can_start_probe)
        self.btn_pause.setEnabled(state in {"running", "paused"})
        self.btn_pause.setText("继续拨测" if state == "paused" else "暂停拨测")
        self.btn_stop.setEnabled(state in {"running", "paused"})

    def _update_thread_hint(self) -> None:
        self.thread_hint.setText(f"当前并发线程上限: {self.concurrency_spin.value()}")

    def _on_concurrency_changed(self, value: int) -> None:
        if self.worker is not None and self._run_state in {"running", "paused"}:
            self.worker.update_concurrency(value)

    def _show_log_context_menu(self, pos: QtCore.QPoint) -> None:
        item = self.preview_table.itemAt(pos)
        if item is None:
            return
        row = item.row()
        url_item = self.preview_table.item(row, 1)
        if url_item is None:
            return
        raw_url = (url_item.text() or "").strip()
        if not raw_url:
            return
        url = normalize_url(raw_url)

        menu = QtWidgets.QMenu(self)
        copy_action = menu.addAction("复制 URL")
        open_action = menu.addAction("打开 URL")
        chosen: QtGui.QAction | None = menu.exec(self.preview_table.viewport().mapToGlobal(pos))  # type: ignore
        if chosen == copy_action:  # type: ignore
            QtWidgets.QApplication.clipboard().setText(url)
        elif chosen == open_action:  # type: ignore
            webbrowser.open(url, new=2, autoraise=True)

    def _on_pause_resume(self) -> None:
        if self.worker is None:
            return
        if self._run_state == "running":
            self.worker.pause_probe()
        elif self._run_state == "paused":
            self.worker.resume_probe()

    def _on_stop(self) -> None:
        if self.worker is None:
            return
        self.worker.stop_probe()
        self.progress_text.setText("正在中止拨测，请稍候...")

    def _on_start(self) -> None:
        if self.workbook is None:
            _: object = QtWidgets.QMessageBox.warning(self, "提示", "请先输入文件路径并等待自动读取结构。")
            return
        if self.column_combo.currentData() is None:
            _: object = QtWidgets.QMessageBox.warning(self, "提示", "请选择网址列。")
            return
        if not self._url_column_valid:
            _: object = QtWidgets.QMessageBox.warning(self, "提示", "当前列未检测到合法网址/域名，请重新选择包含网址/域名的列。")
            return
        output_path = self.output_edit.text().strip()
        if not output_path:
            _: object = QtWidgets.QMessageBox.warning(self, "提示", "请设置输出文件路径。")
            return
        cfg = ProbeConfig(
            file_path=self.file_edit.text().strip(),
            use_office_fallback=self.office_fallback.isChecked(),
            sheet_name=self.sheet_combo.currentText(),
            has_header=self.has_header.isChecked(),
            url_col_idx=int(self.column_combo.currentData()),  # type: ignore
            concurrency=self.concurrency_spin.value(),
            timeout_seconds=self.timeout_spin.value(),
            retry_count=self.retry_spin.value(),
            success_mode=self.success_mode_combo.currentText(),
            custom_codes_input=self.custom_codes_edit.text(),
            output_path=output_path,
            password=self._file_password,
            probe_mode="browser" if self.probe_mode_combo.currentIndex() == 1 else "lightweight",
            extract_title=self.extract_title_check.isChecked(),
            track_redirects=self.track_redirects_check.isChecked(),
        )
        self.progress.setValue(0)
        self.progress_text.setText("进度: 0/0，失败: 0")
        self.metric_label.setText("可访问比例: 0.00%")
        self._log_entries.clear()
        self._mark_run_started()
        self.preview_table.setRowCount(0)
        self._apply_ui_state("running")
        self.worker = ProbeWorker(cfg)
        _: object = self.worker.progress.connect(self._on_progress)
        _: object = self.worker.success.connect(self._on_success)
        _: object = self.worker.aborted.connect(self._on_aborted)
        _: object = self.worker.failed.connect(self._on_failed)
        _: object = self.worker.state_changed.connect(self._on_worker_state_changed)
        _: object = self.worker.finished.connect(self._on_worker_finished)
        self.worker.start()

    def _on_progress(self, completed: int, total: int, failed: int, active_workers: int, result_obj: object) -> None:
        ratio = int((completed / total) * 100) if total else 0
        self.progress.setValue(ratio)
        self.progress_text.setText(
            f"进度: {completed}/{total}，失败: {failed}，执行线程: {active_workers}/{self.concurrency_spin.value()}"
        )
        if completed:
            self.metric_label.setText(f"实时成功率: {((completed - failed) / completed):.2%}")
        if isinstance(result_obj, ProbeResult):
            self._log_entries.append(result_obj)
            if len(self._log_entries) > MAX_LIVE_LOG_ROWS:
                self._log_entries.pop(0)
            self._refresh_log_table()

    def _on_worker_state_changed(self, state: str) -> None:
        if state == "paused":
            self._apply_ui_state("paused")
            base_text = self.progress_text.text().split("（已暂停", 1)[0]
            self.progress_text.setText(f"{base_text}（已暂停，可调整并发数后继续）")
        elif state == "running" and self._run_state == "paused":
            self._apply_ui_state("running")
            base_text = self.progress_text.text().split("（已暂停", 1)[0]
            self.progress_text.setText(base_text)

    def _on_success(self, success_count: int, total_count: int, output_path: str) -> None:
        ratio = (success_count / total_count) if total_count else 0
        self.progress.setValue(100)
        self.progress_text.setText(f"进度: {total_count}/{total_count}，失败: {total_count - success_count}，执行线程: 0/{self.concurrency_spin.value()}")
        self.metric_label.setText(f"可访问比例: {ratio:.2%} ({success_count}/{total_count})")
        self._update_open_output_button_state()
        self._mark_run_finished()
        self._apply_ui_state("idle")
        _: object = QtWidgets.QMessageBox.information(self, "完成", f"拨测完成，结果已写入:\n{output_path}")

    def _on_aborted(self, completed_count: int, total_count: int, failed_count: int) -> None:
        self._apply_ui_state("idle")
        self._mark_run_finished()
        self.progress_text.setText(f"已中止: {completed_count}/{total_count}，失败: {failed_count}，执行线程: 0/{self.concurrency_spin.value()}")
        _: object = QtWidgets.QMessageBox.information(self, "已中止", "拨测任务已中止，你现在可以修改全部配置并重新开始。")

    def _on_failed(self, message: str) -> None:
        self._apply_ui_state("idle")
        self._mark_run_finished()
        _: object = QtWidgets.QMessageBox.critical(self, "失败", message)

    def _on_worker_finished(self) -> None:
        self.worker = None


def main() -> None:
    if os.name == "nt":
        try:
            import ctypes

            _: object = ctypes.windll.shell32.SetCurrentProcessExplicitAppUserModelID("SyntheticMonitoringTool.App")
        except Exception:
            pass
    app = QtWidgets.QApplication(sys.argv)
    app.setStyle("Fusion")
    icon_path = get_app_icon_path()
    if icon_path:
        app.setWindowIcon(QtGui.QIcon(icon_path))
    window = MainWindow()
    if icon_path:
        window.setWindowIcon(QtGui.QIcon(icon_path))
    window.show()
    raise SystemExit(app.exec())


if __name__ == "__main__":
    main()
